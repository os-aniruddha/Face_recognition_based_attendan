import cv2
import face_recognition
import os
from openpyxl import Workbook, load_workbook
import datetime

# Create a folder to store the images
if not os.path.exists("images"):
    os.makedirs("images")

# Function to capture and save images
def capture_and_save_image(person_name):
    # Open the laptop camera
    camera = cv2.VideoCapture(0)

    # Capture a frame
    ret, frame = camera.read()

    # Generate a unique filename (e.g., using timestamp)
    image_name = f"images/{person_name}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"

    # Save the captured image
    cv2.imwrite(image_name, frame)

    # Release the camera
    camera.release()
    cv2.destroyAllWindows()

    return image_name

# Function to recognize faces in the captured image
def recognize_faces(image_path, known_faces):
    # Load the image and convert it to RGB
    image = face_recognition.load_image_file(image_path)
    image_face_encodings = face_recognition.face_encodings(image)

    if not image_face_encodings:
        return None

    for image_face_encoding in image_face_encodings:
        # Compare the face encoding with known faces
        matches = face_recognition.compare_faces(known_faces, image_face_encoding)
        return matches

def main():
    known_faces = []
    attendance_data = {}

    # Create or load an Excel workbook
    if os.path.exists("attendance.xlsx"):
        workbook = load_workbook("attendance.xlsx")
    else:
        workbook = Workbook()
        workbook.active.append(["Name", "Status", "Date"])

    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        known_faces.append(face_recognition.load_image_file(f"images/{row[0]}.jpg"))

    while True:
        try:
            person_name = input("Enter person's name (or 'q' to quit): ")
            if person_name.lower() == 'q':
                break

            image_path = capture_and_save_image(person_name)
            if image_path:
                matches = recognize_faces(image_path, known_faces)

                if matches:
                    matched_name = None
                    for i, match in enumerate(matches):
                        if match:
                            matched_name = sheet.cell(row=i + 2, column=1).value
                            break

                    if matched_name:
                        attendance_data[matched_name] = "Present"
                        sheet.append([matched_name, "Present", datetime.datetime.now()])
                        workbook.save("attendance.xlsx")
                        print(f"{matched_name} is marked as present.")
                    else:
                        print("No matching person found.")
                else:
                    print("No faces detected in the image.")
            else:
                print("Image capture failed.")

        except KeyboardInterrupt:
            break

    print("Attendance recording completed.")
    workbook.close()

if __name__ == "__main__":
    main()
